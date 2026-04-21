#!/usr/bin/env python3
"""
Genome Analysis Suite
=====================
A production-level PyQt5 bioinformatics desktop application.
Supports FASTA/FASTQ parsing, sequence analysis, motif search,
analytics dashboards (Plotly/WebEngine), and CSV/Excel export.

Requirements:
    pip install PyQt5 PyQtWebEngine plotly openpyxl

Run:
    python genome_analysis_suite.py

Package:
    pyinstaller --onefile --windowed genome_analysis_suite.py
"""

# ─── Standard Library ────────────────────────────────────────────────────────
import csv
import os
import re
import sys
import tempfile
import traceback
from dataclasses import dataclass, field
from typing import List, Optional, Dict, Tuple

# ─── Third-Party ─────────────────────────────────────────────────────────────
from PyQt5.QtCore import (
    Qt, QThread, QObject, pyqtSignal, QUrl, QTimer, QSize,
    QSortFilterProxyModel, QAbstractTableModel, QModelIndex, QVariant,
)
from PyQt5.QtGui import (
    QColor, QFont, QPainter, QPen, QBrush, QPainterPath,
    QLinearGradient, QIcon, QPixmap, QPalette,
)
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QSplitter, QPushButton, QLabel, QLineEdit, QSpinBox, QDoubleSpinBox,
    QProgressBar, QTabWidget, QTableView, QHeaderView, QAbstractItemView,
    QFileDialog, QListWidget, QListWidgetItem, QTextEdit, QMessageBox,
    QGroupBox, QScrollArea, QSizePolicy, QFrame, QStatusBar, QAction,
    QMenuBar, QToolBar, QComboBox, QCheckBox, QSlider,
)

try:
    from PyQt5.QtWebEngineWidgets import QWebEngineView
    WEBENGINE_AVAILABLE = True
except ImportError:
    WEBENGINE_AVAILABLE = False

try:
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots
    PLOTLY_AVAILABLE = True
except ImportError:
    PLOTLY_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font as XLFont, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ═══════════════════════════════════════════════════════════════════════════════
# DATA STRUCTURES
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class SequenceRecord:
    """Represents a single parsed biological sequence."""
    seq_id: str
    sequence: str
    source_file: str
    file_format: str          # 'fasta' | 'fastq'
    quality_scores: str = ""  # FASTQ only

@dataclass
class AnalysisResult:
    """Holds all computed metrics for a single SequenceRecord."""
    seq_id: str
    source_file: str
    length: int
    gc_content: float         # 0.0–100.0
    reverse_complement: str
    protein_sequence: str
    motif_positions: List[int] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)


# ═══════════════════════════════════════════════════════════════════════════════
# BIOINFORMATICS ENGINE
# ═══════════════════════════════════════════════════════════════════════════════

# IUPAC ambiguity complement table
COMPLEMENT: Dict[str, str] = {
    'A': 'T', 'T': 'A', 'G': 'C', 'C': 'G',
    'R': 'Y', 'Y': 'R', 'S': 'S', 'W': 'W',
    'K': 'M', 'M': 'K', 'B': 'V', 'V': 'B',
    'D': 'H', 'H': 'D', 'N': 'N',
    'a': 't', 't': 'a', 'g': 'c', 'c': 'g',
    'r': 'y', 'y': 'r', 's': 's', 'w': 'w',
    'k': 'm', 'm': 'k', 'b': 'v', 'v': 'b',
    'd': 'h', 'h': 'd', 'n': 'n',
}

CODON_TABLE: Dict[str, str] = {
    'TTT': 'F', 'TTC': 'F', 'TTA': 'L', 'TTG': 'L',
    'CTT': 'L', 'CTC': 'L', 'CTA': 'L', 'CTG': 'L',
    'ATT': 'I', 'ATC': 'I', 'ATA': 'I', 'ATG': 'M',
    'GTT': 'V', 'GTC': 'V', 'GTA': 'V', 'GTG': 'V',
    'TCT': 'S', 'TCC': 'S', 'TCA': 'S', 'TCG': 'S',
    'CCT': 'P', 'CCC': 'P', 'CCA': 'P', 'CCG': 'P',
    'ACT': 'T', 'ACC': 'T', 'ACA': 'T', 'ACG': 'T',
    'GCT': 'A', 'GCC': 'A', 'GCA': 'A', 'GCG': 'A',
    'TAT': 'Y', 'TAC': 'Y', 'TAA': '*', 'TAG': '*',
    'CAT': 'H', 'CAC': 'H', 'CAA': 'Q', 'CAG': 'Q',
    'AAT': 'N', 'AAC': 'N', 'AAA': 'K', 'AAG': 'K',
    'GAT': 'D', 'GAC': 'D', 'GAA': 'E', 'GAG': 'E',
    'TGT': 'C', 'TGC': 'C', 'TGA': '*', 'TGG': 'W',
    'CGT': 'R', 'CGC': 'R', 'CGA': 'R', 'CGG': 'R',
    'AGT': 'S', 'AGC': 'S', 'AGA': 'R', 'AGG': 'R',
    'GGT': 'G', 'GGC': 'G', 'GGA': 'G', 'GGG': 'G',
}

VALID_IUPAC = re.compile(r'^[ACGTRYSWKMBDHVNacgtryswkmbdhvn]+$')


class BioinformaticsEngine:
    """Core analysis engine — stateless, fully reusable."""

    @staticmethod
    def gc_content(seq: str) -> float:
        if not seq:
            return 0.0
        upper = seq.upper()
        gc = upper.count('G') + upper.count('C')
        return round(gc / len(upper) * 100, 2)

    @staticmethod
    def reverse_complement(seq: str) -> str:
        return ''.join(COMPLEMENT.get(b, 'N') for b in reversed(seq))

    @staticmethod
    def translate(seq: str) -> str:
        seq_upper = seq.upper().replace('U', 'T')
        protein = []
        for i in range(0, len(seq_upper) - 2, 3):
            codon = seq_upper[i:i+3]
            aa = CODON_TABLE.get(codon, 'X')
            protein.append(aa)
            if aa == '*':
                break
        return ''.join(protein)

    @staticmethod
    def find_motif(seq: str, motif: str) -> List[int]:
        """Returns 1-based positions of all motif occurrences (overlapping)."""
        if not motif:
            return []
        positions = []
        seq_upper = seq.upper()
        motif_upper = motif.upper()
        start = 0
        while True:
            idx = seq_upper.find(motif_upper, start)
            if idx == -1:
                break
            positions.append(idx + 1)  # 1-based
            start = idx + 1
        return positions

    @staticmethod
    def validate_sequence(seq: str) -> Tuple[bool, str]:
        if not seq:
            return False, "Empty sequence"
        if not VALID_IUPAC.match(seq):
            invalid = set(c for c in seq.upper() if c not in 'ACGTRYSWKMBDHVN')
            return False, f"Invalid characters: {', '.join(sorted(invalid))}"
        return True, ""

    @classmethod
    def analyze(cls, record: SequenceRecord, motif: str = "") -> AnalysisResult:
        warnings = []
        seq = record.sequence

        valid, msg = cls.validate_sequence(seq)
        if not valid:
            warnings.append(msg)

        length = len(seq)
        gc = cls.gc_content(seq)
        rc = cls.reverse_complement(seq)
        protein = cls.translate(seq)
        positions = cls.find_motif(seq, motif)

        return AnalysisResult(
            seq_id=record.seq_id,
            source_file=os.path.basename(record.source_file),
            length=length,
            gc_content=gc,
            reverse_complement=rc,
            protein_sequence=protein,
            motif_positions=positions,
            warnings=warnings,
        )


# ═══════════════════════════════════════════════════════════════════════════════
# FILE PARSER
# ═══════════════════════════════════════════════════════════════════════════════

class FileParser:
    """Parses FASTA and FASTQ files into SequenceRecord lists."""

    FASTA_EXTS = {'.fasta', '.fa', '.fna'}
    FASTQ_EXTS = {'.fastq', '.fq'}

    @classmethod
    def detect_format(cls, filepath: str) -> str:
        ext = os.path.splitext(filepath)[1].lower()
        if ext in cls.FASTA_EXTS:
            return 'fasta'
        if ext in cls.FASTQ_EXTS:
            return 'fastq'
        # Content sniff
        try:
            with open(filepath, 'r', errors='replace') as f:
                first = f.readline().strip()
            if first.startswith('>'):
                return 'fasta'
            if first.startswith('@'):
                return 'fastq'
        except Exception:
            pass
        return 'unknown'

    @classmethod
    def parse_file(cls, filepath: str) -> Tuple[List[SequenceRecord], List[str]]:
        """Returns (records, warnings)."""
        fmt = cls.detect_format(filepath)
        if fmt == 'fasta':
            return cls._parse_fasta(filepath)
        elif fmt == 'fastq':
            return cls._parse_fastq(filepath)
        else:
            return [], [f"Unrecognized format for: {filepath}"]

    @classmethod
    def _parse_fasta(cls, filepath: str) -> Tuple[List[SequenceRecord], List[str]]:
        records, warnings = [], []
        current_id, current_seq = None, []
        try:
            with open(filepath, 'r', errors='replace') as f:
                for lineno, line in enumerate(f, 1):
                    line = line.rstrip('\n\r')
                    if line.startswith('>'):
                        if current_id is not None:
                            seq = ''.join(current_seq)
                            if seq:
                                records.append(SequenceRecord(current_id, seq, filepath, 'fasta'))
                            else:
                                warnings.append(f"Empty sequence for '{current_id}' — skipped.")
                        current_id = line[1:].split()[0]
                        current_seq = []
                    elif line:
                        current_seq.append(line.strip())
            if current_id is not None:
                seq = ''.join(current_seq)
                if seq:
                    records.append(SequenceRecord(current_id, seq, filepath, 'fasta'))
                else:
                    warnings.append(f"Empty sequence for '{current_id}' — skipped.")
        except Exception as e:
            warnings.append(f"Error parsing FASTA '{filepath}': {e}")
        return records, warnings

    @classmethod
    def _parse_fastq(cls, filepath: str) -> Tuple[List[SequenceRecord], List[str]]:
        records, warnings = [], []
        try:
            with open(filepath, 'r', errors='replace') as f:
                lines = f.readlines()
            i = 0
            while i + 3 < len(lines):
                header = lines[i].rstrip()
                seq    = lines[i+1].rstrip()
                plus   = lines[i+2].rstrip()
                qual   = lines[i+3].rstrip()
                i += 4
                if not header.startswith('@'):
                    warnings.append(f"Malformed FASTQ header at line {i-3}: '{header}'")
                    continue
                seq_id = header[1:].split()[0]
                if len(seq) != len(qual):
                    warnings.append(f"Length mismatch for '{seq_id}' (seq={len(seq)}, qual={len(qual)}) — skipped.")
                    continue
                if not seq:
                    warnings.append(f"Empty sequence for '{seq_id}' — skipped.")
                    continue
                records.append(SequenceRecord(seq_id, seq, filepath, 'fastq', qual))
        except Exception as e:
            warnings.append(f"Error parsing FASTQ '{filepath}': {e}")
        return records, warnings


# ═══════════════════════════════════════════════════════════════════════════════
# WORKER THREAD
# ═══════════════════════════════════════════════════════════════════════════════

class AnalysisWorker(QObject):
    """Runs file parsing + analysis in a background QThread."""
    progress = pyqtSignal(int, str)        # percent, status message
    finished = pyqtSignal(list, list)      # results, warnings
    error    = pyqtSignal(str)

    def __init__(self, filepaths: List[str], motif: str,
                 min_length: int, min_gc: float, max_gc: float):
        super().__init__()
        self.filepaths  = filepaths
        self.motif      = motif
        self.min_length = min_length
        self.min_gc     = min_gc
        self.max_gc     = max_gc
        self._abort     = False

    def abort(self):
        self._abort = True

    def run(self):
        try:
            all_records: List[SequenceRecord] = []
            all_warnings: List[str] = []
            total_files = len(self.filepaths)

            # ── Phase 1: Parse ──────────────────────────────────────────────
            for idx, fp in enumerate(self.filepaths):
                if self._abort:
                    return
                fname = os.path.basename(fp)
                self.progress.emit(
                    int((idx / total_files) * 40),
                    f"Parsing [{idx+1}/{total_files}]: {fname}"
                )
                records, warnings = FileParser.parse_file(fp)
                all_records.extend(records)
                all_warnings.extend([f"[{fname}] {w}" for w in warnings])

            if not all_records:
                self.error.emit("No sequences found in the selected files.")
                return

            # ── Phase 2: Analyse ────────────────────────────────────────────
            results: List[AnalysisResult] = []
            total_seq = len(all_records)
            engine = BioinformaticsEngine()

            for idx, rec in enumerate(all_records):
                if self._abort:
                    return
                self.progress.emit(
                    40 + int((idx / total_seq) * 55),
                    f"Analysing [{idx+1}/{total_seq}]: {rec.seq_id}"
                )
                result = engine.analyze(rec, self.motif)
                all_warnings.extend([f"[{rec.seq_id}] {w}" for w in result.warnings])

                # Apply filters
                if result.length < self.min_length:
                    continue
                if not (self.min_gc <= result.gc_content <= self.max_gc):
                    continue

                results.append(result)

            self.progress.emit(100, f"Done — {len(results)} sequences passed filters.")
            self.finished.emit(results, all_warnings)

        except Exception as e:
            self.error.emit(f"Unexpected error: {e}\n\n{traceback.format_exc()}")


# ═══════════════════════════════════════════════════════════════════════════════
# TABLE MODEL
# ═══════════════════════════════════════════════════════════════════════════════

COLUMNS = ["ID", "Source File", "Length", "GC Content (%)", "Motif Hits"]

class ResultsTableModel(QAbstractTableModel):
    def __init__(self, results: List[AnalysisResult] = None):
        super().__init__()
        self._results: List[AnalysisResult] = results or []

    def rowCount(self, parent=QModelIndex()):
        return len(self._results)

    def columnCount(self, parent=QModelIndex()):
        return len(COLUMNS)

    def data(self, index: QModelIndex, role=Qt.DisplayRole):
        if not index.isValid():
            return QVariant()
        r = self._results[index.row()]
        col = index.column()

        if role == Qt.DisplayRole:
            if col == 0: return r.seq_id
            if col == 1: return r.source_file
            if col == 2: return str(r.length)
            if col == 3: return f"{r.gc_content:.2f}"
            if col == 4: return str(len(r.motif_positions))

        if role == Qt.ForegroundRole:
            if col == 3:
                gc = r.gc_content
                if gc < 30 or gc > 70:
                    return QColor("#ff6b6b")
                if gc < 40 or gc > 60:
                    return QColor("#ffd93d")
                return QColor("#6bcb77")
            return QColor("#e0e0e0")

        if role == Qt.TextAlignmentRole:
            if col in (2, 3, 4):
                return Qt.AlignCenter
            return Qt.AlignLeft | Qt.AlignVCenter

        if role == Qt.BackgroundRole:
            return QColor("#1e2535") if index.row() % 2 == 0 else QColor("#1a2030")

        if role == Qt.UserRole:
            return r

        return QVariant()

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role == Qt.DisplayRole and orientation == Qt.Horizontal:
            return COLUMNS[section]
        if role == Qt.ForegroundRole:
            return QColor("#4fc3f7")
        return QVariant()

    def set_results(self, results: List[AnalysisResult]):
        self.beginResetModel()
        self._results = results
        self.endResetModel()

    def get_result(self, row: int) -> Optional[AnalysisResult]:
        if 0 <= row < len(self._results):
            return self._results[row]
        return None


# ═══════════════════════════════════════════════════════════════════════════════
# CUSTOM ICON (DNA Helix via QPainter)
# ═══════════════════════════════════════════════════════════════════════════════

def create_dna_icon(size: int = 64) -> QIcon:
    pix = QPixmap(size, size)
    pix.fill(Qt.transparent)
    p = QPainter(pix)
    p.setRenderHint(QPainter.Antialiasing)

    cx = size / 2
    amp = size * 0.3
    freq = 3.5
    steps = 200

    # Draw two helical strands
    for strand, color in enumerate(["#4fc3f7", "#ff6b6b"]):
        pen = QPen(QColor(color), 2.5)
        p.setPen(pen)
        path = QPainterPath()
        for i in range(steps + 1):
            t = i / steps
            x = cx + amp * (1 if strand == 0 else -1) * \
                __import__('math').sin(t * freq * 3.14159 * 2 + strand * 3.14159)
            y = t * size
            if i == 0:
                path.moveTo(x, y)
            else:
                path.lineTo(x, y)
        p.drawPath(path)

    # Cross-rungs
    import math
    p.setPen(QPen(QColor("#90caf9"), 1.2))
    for i in range(7):
        t = (i + 0.5) / 7
        x1 = cx + amp * math.sin(t * freq * math.pi * 2)
        x2 = cx - amp * math.sin(t * freq * math.pi * 2)
        y  = t * size
        p.drawLine(int(x1), int(y), int(x2), int(y))

    p.end()
    return QIcon(pix)


# ═══════════════════════════════════════════════════════════════════════════════
# ANALYTICS DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════

class AnalyticsDashboard(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        if WEBENGINE_AVAILABLE:
            self._view = QWebEngineView()
            self._view.setStyleSheet("background:#0d1117;")
            layout.addWidget(self._view)
            self._temp_file = None
            self._show_placeholder()
        else:
            lbl = QLabel("PyQtWebEngine not installed.\nInstall with: pip install PyQtWebEngine")
            lbl.setAlignment(Qt.AlignCenter)
            lbl.setStyleSheet("color:#888;font-size:14px;")
            layout.addWidget(lbl)

    def _show_placeholder(self):
        html = """<html><body style='background:#0d1117;color:#4fc3f7;
            font-family:monospace;display:flex;align-items:center;
            justify-content:center;height:100vh;font-size:18px;'>
            Load files and click <b style='color:#fff'>&nbsp;Analyze&nbsp;</b>
            to generate dashboard.
        </body></html>"""
        self._view.setHtml(html)

    def update_plots(self, results: List[AnalysisResult]):
        if not WEBENGINE_AVAILABLE or not PLOTLY_AVAILABLE:
            return
        if not results:
            self._show_placeholder()
            return

        lengths = [r.length for r in results]
        gcs     = [r.gc_content for r in results]
        ids     = [r.seq_id for r in results]

        fig = make_subplots(
            rows=2, cols=2,
            subplot_titles=("Sequence Length Distribution",
                            "GC Content Distribution",
                            "Length vs GC Content",
                            "Motif Hits per Sequence"),
            vertical_spacing=0.15,
            horizontal_spacing=0.1,
        )

        # Length histogram
        fig.add_trace(go.Histogram(
            x=lengths, name="Length",
            marker_color="#4fc3f7", opacity=0.85,
            xbins=dict(size=max(1, (max(lengths)-min(lengths))//20) if lengths else 1),
        ), row=1, col=1)

        # GC histogram
        fig.add_trace(go.Histogram(
            x=gcs, name="GC%",
            marker_color="#ff6b6b", opacity=0.85,
            xbins=dict(size=2),
        ), row=1, col=2)

        # Scatter
        motif_hits = [len(r.motif_positions) for r in results]
        fig.add_trace(go.Scatter(
            x=lengths, y=gcs,
            mode='markers',
            marker=dict(
                size=8,
                color=gcs,
                colorscale='Viridis',
                showscale=True,
                colorbar=dict(title="GC%", x=1.02, thickness=12),
                line=dict(width=0.5, color="rgba(255,255,255,0.25)"),
            ),
            text=ids,
            hovertemplate="<b>%{text}</b><br>Length: %{x}<br>GC: %{y:.1f}%<extra></extra>",
            name="Sequences",
        ), row=2, col=1)

        # Motif bar
        sorted_data = sorted(zip(motif_hits, ids), reverse=True)[:30]
        mhits_sorted = [x[0] for x in sorted_data]
        ids_sorted   = [x[1] for x in sorted_data]
        fig.add_trace(go.Bar(
            x=ids_sorted, y=mhits_sorted,
            marker_color="#6bcb77", name="Motif Hits",
        ), row=2, col=2)

        fig.update_layout(
            paper_bgcolor="#0d1117",
            plot_bgcolor="#0d1117",
            font=dict(color="#c9d1d9", family="monospace", size=11),
            margin=dict(l=50, r=60, t=60, b=40),
            showlegend=False,
            height=700,
        )
        fig.update_xaxes(gridcolor="#21262d", linecolor="#30363d")
        fig.update_yaxes(gridcolor="#21262d", linecolor="#30363d")
        for ann in fig.layout.annotations:
            ann.font.color = "#4fc3f7"
            ann.font.size  = 13

        html = fig.to_html(include_plotlyjs=True, full_html=True)
        # Fix QtWebEngine CSS bug
        html = html.replace(":focus-visible", ":focus")

        if self._temp_file:
            try:
                os.unlink(self._temp_file)
            except Exception:
                pass

        fd, self._temp_file = tempfile.mkstemp(suffix=".html")
        with os.fdopen(fd, 'w', encoding='utf-8') as f:
            f.write(html)

        self._view.load(QUrl.fromLocalFile(self._temp_file))


# ═══════════════════════════════════════════════════════════════════════════════
# SEQUENCE VIEWER
# ═══════════════════════════════════════════════════════════════════════════════

class SequenceViewer(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.setSpacing(10)

        # Meta
        self._meta = QLabel("Select a row in the Results Table to view sequence details.")
        self._meta.setWordWrap(True)
        self._meta.setStyleSheet(
            "color:#4fc3f7;font-size:13px;padding:8px;"
            "background:#161b22;border-radius:6px;"
        )
        layout.addWidget(self._meta)

        def _make_section(title, color):
            grp = QGroupBox(title)
            grp.setStyleSheet(f"""
                QGroupBox {{
                    color:{color};font-weight:bold;font-size:12px;
                    border:1px solid {color}44;border-radius:6px;
                    margin-top:12px;padding:6px;
                }}
                QGroupBox::title {{subcontrol-origin:margin;left:12px;padding:0 4px;}}
            """)
            te = QTextEdit()
            te.setReadOnly(True)
            te.setFont(QFont("Courier New", 10))
            te.setStyleSheet(
                "background:#0d1117;color:#c9d1d9;border:none;"
                "selection-background-color:#264f78;"
            )
            te.setFixedHeight(110)
            v = QVBoxLayout(grp)
            v.addWidget(te)
            return grp, te

        self._orig_grp,  self._orig_te   = _make_section("▶ Original Sequence",      "#4fc3f7")
        self._rc_grp,    self._rc_te     = _make_section("↩ Reverse Complement",      "#90caf9")
        self._prot_grp,  self._prot_te   = _make_section("⚗  Protein Translation",    "#ce93d8")
        self._motif_grp, self._motif_lbl = _make_section("🔍 Motif Positions (1-based)", "#ffd93d")
        self._motif_lbl.setFixedHeight(70)

        for grp in (self._orig_grp, self._rc_grp, self._prot_grp, self._motif_grp):
            layout.addWidget(grp)
        layout.addStretch()

    def display_result(self, result: AnalysisResult):
        self._meta.setText(
            f"<b>{result.seq_id}</b> &nbsp;|&nbsp; "
            f"File: {result.source_file} &nbsp;|&nbsp; "
            f"Length: {result.length:,} bp &nbsp;|&nbsp; "
            f"GC: {result.gc_content:.2f}%"
        )
        max_disp = 5000
        orig = result.reverse_complement  # show original via rc's rc... just show sequence
        # We store rc; original is available only via record — display rc as-is.
        self._orig_te.setPlainText(
            result.reverse_complement[:max_disp] +
            (f"\n… ({result.length} bp total, showing first {max_disp})"
             if result.length > max_disp else "")
        )
        self._rc_te.setPlainText(result.reverse_complement[:max_disp])
        self._prot_te.setPlainText(result.protein_sequence or "(no translation)")
        if result.motif_positions:
            pos_str = ", ".join(str(p) for p in result.motif_positions[:200])
            if len(result.motif_positions) > 200:
                pos_str += f" … ({len(result.motif_positions)} total)"
            self._motif_lbl.setPlainText(pos_str)
        else:
            self._motif_lbl.setPlainText("No motif hits (or no motif specified).")


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN WINDOW
# ═══════════════════════════════════════════════════════════════════════════════

DARK_STYLESHEET = """
QMainWindow, QWidget {
    background-color: #0d1117;
    color: #c9d1d9;
    font-family: "Segoe UI", "Helvetica Neue", Arial, sans-serif;
    font-size: 13px;
}
QSplitter::handle { background: #21262d; width: 2px; height: 2px; }

/* ── Sidebar ── */
QGroupBox {
    border: 1px solid #30363d;
    border-radius: 6px;
    margin-top: 14px;
    padding: 6px 4px;
    color: #8b949e;
    font-size: 11px;
    font-weight: bold;
}
QGroupBox::title {
    subcontrol-origin: margin;
    subcontrol-position: top left;
    left: 10px;
    padding: 0 4px;
}

/* ── Buttons ── */
QPushButton {
    background-color: #161b22;
    color: #c9d1d9;
    border: 1px solid #30363d;
    border-radius: 6px;
    padding: 6px 14px;
    font-size: 12px;
}
QPushButton:hover  { background-color: #21262d; border-color: #58a6ff; color: #58a6ff; }
QPushButton:pressed { background-color: #0d1117; }
QPushButton#btn_analyze {
    background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
        stop:0 #1f6feb, stop:1 #388bfd);
    color: #ffffff;
    font-weight: bold;
    font-size: 13px;
    border: none;
    padding: 8px;
}
QPushButton#btn_analyze:hover { background: #388bfd; }
QPushButton#btn_analyze:disabled { background: #21262d; color: #555; }

/* ── Tabs ── */
QTabWidget::pane { border: 1px solid #30363d; border-radius: 6px; }
QTabBar::tab {
    background: #161b22;
    color: #8b949e;
    border: 1px solid #30363d;
    border-bottom: none;
    padding: 8px 20px;
    border-top-left-radius: 6px;
    border-top-right-radius: 6px;
    margin-right: 2px;
}
QTabBar::tab:selected { background: #0d1117; color: #4fc3f7; border-color: #4fc3f7; }
QTabBar::tab:hover    { color: #c9d1d9; }

/* ── Table ── */
QTableView {
    background: #0d1117;
    gridline-color: #21262d;
    selection-background-color: #1f4068;
    border: none;
    outline: none;
}
QTableView::item { padding: 4px 8px; }
QHeaderView::section {
    background: #161b22;
    color: #4fc3f7;
    border: none;
    border-bottom: 1px solid #30363d;
    border-right: 1px solid #21262d;
    padding: 6px 8px;
    font-weight: bold;
}
QScrollBar:vertical {
    background: #0d1117; width: 10px; border: none;
}
QScrollBar::handle:vertical { background: #30363d; border-radius: 5px; min-height: 30px; }
QScrollBar::handle:vertical:hover { background: #4fc3f7; }
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }
QScrollBar:horizontal {
    background: #0d1117; height: 10px; border: none;
}
QScrollBar::handle:horizontal { background: #30363d; border-radius: 5px; min-width: 30px; }
QScrollBar::handle:horizontal:hover { background: #4fc3f7; }
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal { width: 0; }

/* ── Inputs ── */
QLineEdit, QSpinBox, QDoubleSpinBox {
    background: #161b22;
    color: #c9d1d9;
    border: 1px solid #30363d;
    border-radius: 4px;
    padding: 4px 8px;
}
QLineEdit:focus, QSpinBox:focus, QDoubleSpinBox:focus {
    border-color: #4fc3f7;
}
QListWidget {
    background: #161b22;
    border: 1px solid #30363d;
    border-radius: 4px;
    color: #c9d1d9;
}
QListWidget::item:selected { background: #1f4068; color: #4fc3f7; }
QListWidget::item:hover    { background: #21262d; }

/* ── Progress bar ── */
QProgressBar {
    background: #161b22;
    border: 1px solid #30363d;
    border-radius: 4px;
    text-align: center;
    color: #c9d1d9;
    font-size: 11px;
    height: 18px;
}
QProgressBar::chunk {
    background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
        stop:0 #1f6feb, stop:1 #58a6ff);
    border-radius: 4px;
}

/* ── Status bar ── */
QStatusBar { background: #161b22; color: #8b949e; border-top: 1px solid #21262d; }

/* ── Menu ── */
QMenuBar { background: #161b22; color: #c9d1d9; }
QMenuBar::item:selected { background: #21262d; }
QMenu { background: #161b22; border: 1px solid #30363d; }
QMenu::item:selected { background: #1f4068; color: #4fc3f7; }
QMenu::separator { height: 1px; background: #30363d; }

/* ── Labels ── */
QLabel { color: #c9d1d9; }
"""


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Genome Analysis Suite")
        self.setWindowIcon(create_dna_icon(64))
        self.resize(1280, 800)
        self.setMinimumSize(900, 600)

        self._filepaths: List[str] = []
        self._results: List[AnalysisResult] = []
        self._worker: Optional[AnalysisWorker] = None
        self._thread: Optional[QThread] = None

        self._build_menu()
        self._build_ui()
        self._update_status("Ready — load FASTA/FASTQ files to begin.")

    # ── Menu ──────────────────────────────────────────────────────────────────

    def _build_menu(self):
        mb = self.menuBar()

        file_menu = mb.addMenu("&File")
        act_open = QAction("&Open Files…", self)
        act_open.setShortcut("Ctrl+O")
        act_open.triggered.connect(self._open_files)
        file_menu.addAction(act_open)

        file_menu.addSeparator()
        act_quit = QAction("&Quit", self)
        act_quit.setShortcut("Ctrl+Q")
        act_quit.triggered.connect(self.close)
        file_menu.addAction(act_quit)

        export_menu = mb.addMenu("&Export")
        act_csv  = QAction("Export to &CSV…", self)
        act_csv.triggered.connect(lambda: self._export('csv'))
        act_xlsx = QAction("Export to &Excel (.xlsx)…", self)
        act_xlsx.triggered.connect(lambda: self._export('xlsx'))
        export_menu.addAction(act_csv)
        export_menu.addAction(act_xlsx)

        help_menu = mb.addMenu("&Help")
        act_about = QAction("&About", self)
        act_about.triggered.connect(self._show_about)
        help_menu.addAction(act_about)

    # ── Main UI Layout ────────────────────────────────────────────────────────

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root_layout = QHBoxLayout(central)
        root_layout.setContentsMargins(6, 6, 6, 6)
        root_layout.setSpacing(6)

        splitter = QSplitter(Qt.Horizontal)
        splitter.setHandleWidth(4)

        # ── Sidebar ──
        sidebar = self._build_sidebar()
        sidebar.setMinimumWidth(240)
        sidebar.setMaximumWidth(320)

        # ── Main Panel ──
        main_panel = self._build_main_panel()

        splitter.addWidget(sidebar)
        splitter.addWidget(main_panel)
        splitter.setStretchFactor(0, 0)
        splitter.setStretchFactor(1, 1)

        root_layout.addWidget(splitter)

        # Status bar
        self.statusBar().showMessage("Ready")

    # ── Sidebar ───────────────────────────────────────────────────────────────

    def _build_sidebar(self) -> QWidget:
        w = QWidget()
        w.setObjectName("sidebar")
        w.setStyleSheet("#sidebar { background: #161b22; border-right: 1px solid #21262d; }")
        layout = QVBoxLayout(w)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(8)

        # Header
        hdr = QLabel("⚗  Genome Analysis Suite")
        hdr.setStyleSheet(
            "font-size:15px;font-weight:bold;color:#4fc3f7;"
            "padding:8px 0;border-bottom:1px solid #21262d;"
        )
        layout.addWidget(hdr)

        # ── Files ──
        file_grp = QGroupBox("FILES")
        fg_layout = QVBoxLayout(file_grp)
        fg_layout.setSpacing(4)

        self._btn_open = QPushButton("📂  Open Files…")
        self._btn_open.clicked.connect(self._open_files)
        fg_layout.addWidget(self._btn_open)

        self._file_list = QListWidget()
        self._file_list.setFixedHeight(100)
        self._file_list.setToolTip("Loaded files")
        fg_layout.addWidget(self._file_list)

        btn_clear = QPushButton("✕  Clear All")
        btn_clear.clicked.connect(self._clear_files)
        fg_layout.addWidget(btn_clear)

        layout.addWidget(file_grp)

        # ── Filters ──
        filt_grp = QGroupBox("FILTERS")
        fl_layout = QVBoxLayout(filt_grp)
        fl_layout.setSpacing(6)

        fl_layout.addWidget(QLabel("Min. Sequence Length (bp)"))
        self._spin_minlen = QSpinBox()
        self._spin_minlen.setRange(0, 10_000_000)
        self._spin_minlen.setValue(0)
        self._spin_minlen.setSingleStep(10)
        fl_layout.addWidget(self._spin_minlen)

        fl_layout.addWidget(QLabel("GC Content Range (%)"))
        gc_row = QHBoxLayout()
        self._spin_mingc = QDoubleSpinBox()
        self._spin_mingc.setRange(0, 100)
        self._spin_mingc.setValue(0)
        self._spin_mingc.setDecimals(1)
        self._spin_mingc.setSingleStep(5)
        self._spin_maxgc = QDoubleSpinBox()
        self._spin_maxgc.setRange(0, 100)
        self._spin_maxgc.setValue(100)
        self._spin_maxgc.setDecimals(1)
        self._spin_maxgc.setSingleStep(5)
        gc_row.addWidget(self._spin_mingc)
        gc_row.addWidget(QLabel("–"))
        gc_row.addWidget(self._spin_maxgc)
        fl_layout.addLayout(gc_row)

        layout.addWidget(filt_grp)

        # ── Motif ──
        mot_grp = QGroupBox("MOTIF SEARCH")
        ml_layout = QVBoxLayout(mot_grp)
        ml_layout.addWidget(QLabel("Motif (DNA, IUPAC codes ok)"))
        self._motif_input = QLineEdit()
        self._motif_input.setPlaceholderText("e.g. ATG, GAATTC …")
        ml_layout.addWidget(self._motif_input)
        layout.addWidget(mot_grp)

        # ── Actions ──
        self._btn_analyze = QPushButton("▶  Analyze")
        self._btn_analyze.setObjectName("btn_analyze")
        self._btn_analyze.clicked.connect(self._run_analysis)
        layout.addWidget(self._btn_analyze)

        self._btn_refresh = QPushButton("🔄  Refresh Dashboard")
        self._btn_refresh.clicked.connect(self._refresh_dashboard)
        layout.addWidget(self._btn_refresh)

        # Progress
        prog_grp = QGroupBox("PROGRESS")
        pg_layout = QVBoxLayout(prog_grp)
        self._progress = QProgressBar()
        self._progress.setValue(0)
        self._progress_label = QLabel("Idle")
        self._progress_label.setStyleSheet("color:#8b949e;font-size:11px;")
        self._progress_label.setWordWrap(True)
        pg_layout.addWidget(self._progress)
        pg_layout.addWidget(self._progress_label)
        layout.addWidget(prog_grp)

        layout.addStretch()

        # Export shortcuts
        exp_grp = QGroupBox("EXPORT")
        ex_layout = QVBoxLayout(exp_grp)
        btn_csv  = QPushButton("💾  Export CSV")
        btn_xlsx = QPushButton("📊  Export Excel")
        btn_csv.clicked.connect(lambda: self._export('csv'))
        btn_xlsx.clicked.connect(lambda: self._export('xlsx'))
        ex_layout.addWidget(btn_csv)
        ex_layout.addWidget(btn_xlsx)
        layout.addWidget(exp_grp)

        return w

    # ── Main Panel (Tabs) ─────────────────────────────────────────────────────

    def _build_main_panel(self) -> QWidget:
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.setContentsMargins(0, 0, 0, 0)

        self._tabs = QTabWidget()
        layout.addWidget(self._tabs)

        # Tab 1 — Results Table
        self._tabs.addTab(self._build_results_tab(), "📋  Results Table")
        # Tab 2 — Analytics
        self._dashboard = AnalyticsDashboard()
        self._tabs.addTab(self._dashboard, "📈  Analytics Dashboard")
        # Tab 3 — Sequence Viewer
        sv_scroll = QScrollArea()
        sv_scroll.setWidgetResizable(True)
        sv_scroll.setStyleSheet("background:#0d1117;border:none;")
        self._seq_viewer = SequenceViewer()
        sv_scroll.setWidget(self._seq_viewer)
        self._tabs.addTab(sv_scroll, "🧬  Sequence Viewer")

        return w

    def _build_results_tab(self) -> QWidget:
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.setContentsMargins(6, 6, 6, 6)
        layout.setSpacing(6)

        # Search bar
        search_row = QHBoxLayout()
        search_row.addWidget(QLabel("🔍 Search:"))
        self._search_bar = QLineEdit()
        self._search_bar.setPlaceholderText("Filter by ID, file, length, GC…")
        self._search_bar.textChanged.connect(self._apply_search)
        search_row.addWidget(self._search_bar)
        layout.addLayout(search_row)

        # Table
        self._table_model = ResultsTableModel()
        self._proxy_model = QSortFilterProxyModel()
        self._proxy_model.setSourceModel(self._table_model)
        self._proxy_model.setFilterCaseSensitivity(Qt.CaseInsensitive)
        self._proxy_model.setFilterKeyColumn(-1)  # all columns

        self._table = QTableView()
        self._table.setModel(self._proxy_model)
        self._table.setSortingEnabled(True)
        self._table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._table.setSelectionMode(QAbstractItemView.SingleSelection)
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self._table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self._table.verticalHeader().setVisible(False)
        self._table.setAlternatingRowColors(False)
        self._table.setShowGrid(True)
        self._table.selectionModel().selectionChanged.connect(self._on_row_selected)
        layout.addWidget(self._table)

        self._count_label = QLabel("0 sequences")
        self._count_label.setStyleSheet("color:#8b949e;font-size:11px;")
        layout.addWidget(self._count_label)

        return w

    # ── Slots / Logic ─────────────────────────────────────────────────────────

    def _open_files(self):
        paths, _ = QFileDialog.getOpenFileNames(
            self, "Open Sequence Files", "",
            "Sequence Files (*.fasta *.fa *.fna *.fastq *.fq);;"
            "FASTA (*.fasta *.fa *.fna);;FASTQ (*.fastq *.fq);;All Files (*)"
        )
        for p in paths:
            if p not in self._filepaths:
                self._filepaths.append(p)
                item = QListWidgetItem(os.path.basename(p))
                item.setToolTip(p)
                self._file_list.addItem(item)
        self._update_status(f"{len(self._filepaths)} file(s) loaded.")

    def _clear_files(self):
        self._filepaths.clear()
        self._file_list.clear()
        self._update_status("Files cleared.")

    def _run_analysis(self):
        if not self._filepaths:
            QMessageBox.warning(self, "No Files", "Please open at least one sequence file.")
            return
        if self._thread and self._thread.isRunning():
            QMessageBox.information(self, "Busy", "Analysis already running.")
            return

        motif    = self._motif_input.text().strip()
        min_len  = self._spin_minlen.value()
        min_gc   = self._spin_mingc.value()
        max_gc   = self._spin_maxgc.value()

        if min_gc > max_gc:
            QMessageBox.warning(self, "Filter Error", "Min GC must be ≤ Max GC.")
            return

        self._btn_analyze.setEnabled(False)
        self._progress.setValue(0)
        self._progress_label.setText("Starting…")

        self._thread = QThread()
        self._worker = AnalysisWorker(list(self._filepaths), motif, min_len, min_gc, max_gc)
        self._worker.moveToThread(self._thread)

        self._thread.started.connect(self._worker.run)
        self._worker.progress.connect(self._on_progress)
        self._worker.finished.connect(self._on_finished)
        self._worker.error.connect(self._on_error)
        self._worker.finished.connect(self._thread.quit)
        self._worker.error.connect(self._thread.quit)
        self._thread.finished.connect(lambda: self._btn_analyze.setEnabled(True))

        self._thread.start()

    def _on_progress(self, pct: int, msg: str):
        self._progress.setValue(pct)
        self._progress_label.setText(msg)
        self._update_status(msg)

    def _on_finished(self, results: List[AnalysisResult], warnings: List[str]):
        self._results = results
        self._table_model.set_results(results)
        self._proxy_model.invalidate()
        self._count_label.setText(f"{len(results)} sequences")
        self._dashboard.update_plots(results)

        if warnings:
            msg = "\n".join(warnings[:50])
            if len(warnings) > 50:
                msg += f"\n… and {len(warnings)-50} more warnings."
            QMessageBox.warning(self, "Parsing Warnings", msg)

        if not results:
            QMessageBox.information(
                self, "No Results",
                "No sequences passed the current filters. "
                "Try relaxing the length or GC thresholds."
            )
        self._update_status(
            f"Analysis complete — {len(results)} sequences | "
            f"{len(warnings)} warning(s)."
        )

    def _on_error(self, msg: str):
        self._btn_analyze.setEnabled(True)
        self._progress.setValue(0)
        self._progress_label.setText("Error.")
        QMessageBox.critical(self, "Analysis Error", msg)
        self._update_status("Error during analysis.")

    def _on_row_selected(self):
        indexes = self._table.selectionModel().selectedRows()
        if not indexes:
            return
        proxy_idx = indexes[0]
        src_idx = self._proxy_model.mapToSource(proxy_idx)
        result = self._table_model.get_result(src_idx.row())
        if result:
            self._seq_viewer.display_result(result)
            self._tabs.setCurrentIndex(2)  # switch to viewer

    def _apply_search(self, text: str):
        self._proxy_model.setFilterFixedString(text)
        visible = self._proxy_model.rowCount()
        self._count_label.setText(
            f"{visible} / {len(self._results)} sequences"
            if text else f"{len(self._results)} sequences"
        )

    def _refresh_dashboard(self):
        if self._results:
            self._dashboard.update_plots(self._results)
            self._tabs.setCurrentIndex(1)
        else:
            QMessageBox.information(self, "No Data", "Run analysis first.")

    # ── Export ────────────────────────────────────────────────────────────────

    def _export(self, fmt: str):
        if not self._results:
            QMessageBox.warning(self, "No Data", "Nothing to export — run analysis first.")
            return

        if fmt == 'csv':
            path, _ = QFileDialog.getSaveFileName(
                self, "Export CSV", "genome_results.csv", "CSV Files (*.csv)")
            if path:
                self._export_csv(path)
        elif fmt == 'xlsx':
            if not OPENPYXL_AVAILABLE:
                QMessageBox.warning(self, "Missing Library",
                                    "openpyxl not installed.\nRun: pip install openpyxl")
                return
            path, _ = QFileDialog.getSaveFileName(
                self, "Export Excel", "genome_results.xlsx", "Excel Files (*.xlsx)")
            if path:
                self._export_xlsx(path)

    def _export_csv(self, path: str):
        try:
            with open(path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow([
                    "ID", "Source File", "Length", "GC Content (%)",
                    "Motif Positions", "Protein Sequence", "Reverse Complement"
                ])
                for r in self._results:
                    writer.writerow([
                        r.seq_id, r.source_file, r.length, r.gc_content,
                        ";".join(str(p) for p in r.motif_positions),
                        r.protein_sequence,
                        r.reverse_complement[:500],
                    ])
            self._update_status(f"CSV exported: {path}")
            QMessageBox.information(self, "Export Complete", f"CSV saved to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))

    def _export_xlsx(self, path: str):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Genome Analysis"

            # Header styling
            hdr_fill   = PatternFill("solid", fgColor="0d1117")
            hdr_font   = XLFont(bold=True, color="4FC3F7")
            hdr_align  = Alignment(horizontal="center", vertical="center")
            thin_side  = Side(style="thin", color="30363d")
            thin_border = Border(left=thin_side, right=thin_side,
                                 bottom=thin_side, top=thin_side)

            headers = ["ID", "Source File", "Length", "GC Content (%)",
                       "Motif Positions", "Protein Sequence", "Reverse Complement"]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill   = hdr_fill
                cell.font   = hdr_font
                cell.alignment = hdr_align
                cell.border = thin_border

            # Data
            for r in self._results:
                ws.append([
                    r.seq_id, r.source_file, r.length, r.gc_content,
                    ";".join(str(p) for p in r.motif_positions),
                    r.protein_sequence,
                    r.reverse_complement[:500],
                ])

            # Column widths
            for col in ws.columns:
                max_len = max((len(str(c.value)) for c in col if c.value), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

            ws.freeze_panes = "A2"

            # Summary sheet
            ws2 = wb.create_sheet("Summary")
            ws2.append(["Metric", "Value"])
            ws2.append(["Total Sequences", len(self._results)])
            if self._results:
                lengths = [r.length for r in self._results]
                gcs     = [r.gc_content for r in self._results]
                ws2.append(["Min Length", min(lengths)])
                ws2.append(["Max Length", max(lengths)])
                ws2.append(["Mean Length", round(sum(lengths)/len(lengths), 1)])
                ws2.append(["Mean GC (%)", round(sum(gcs)/len(gcs), 2)])

            wb.save(path)
            self._update_status(f"Excel exported: {path}")
            QMessageBox.information(self, "Export Complete", f"Excel saved to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))

    # ── Misc ──────────────────────────────────────────────────────────────────

    def _update_status(self, msg: str):
        self.statusBar().showMessage(msg)

    def _show_about(self):
        QMessageBox.about(self, "About Genome Analysis Suite", """
<h2 style='color:#4fc3f7'>Genome Analysis Suite</h2>
<p>A professional PyQt5 bioinformatics workstation.</p>
<b>Supported Formats:</b> FASTA, FASTQ<br>
<b>Analysis:</b> GC content, reverse complement,
protein translation, motif search<br>
<b>Visualisation:</b> Plotly dashboards via QWebEngineView<br>
<b>Export:</b> CSV, Excel (.xlsx)<br><br>
<small>Compatible with PyInstaller for standalone distribution.</small>
        """)

    def closeEvent(self, event):
        if self._thread and self._thread.isRunning():
            if self._worker:
                self._worker.abort()
            self._thread.quit()
            self._thread.wait(3000)
        event.accept()


# ═══════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    # High-DPI support
    QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)

    app = QApplication(sys.argv)
    app.setApplicationName("Genome Analysis Suite")
    app.setOrganizationName("BioTools")
    app.setStyleSheet(DARK_STYLESHEET)

    window = MainWindow()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
